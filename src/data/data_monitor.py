# coding: utf-8

import os
import re
import time
import psutil
import matplotlib.pyplot as plt
from datetime import datetime
import platform
import cpuinfo
import GPUtil
import subprocess
from threading import Thread
import inspect
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from pynvml import (nvmlInit, nvmlDeviceGetHandleByIndex, nvmlDeviceGetUtilizationRates,
                    nvmlDeviceGetMemoryInfo, nvmlShutdown)

from log import formatted_logging
from config.path_helpers import PROD_LOG_PATH

if platform.system() == 'Windows':
    import wmi

    PRODUCTION_SOFTWARE_VERSION_TXT_PATH = "C:\\Program Files\\UnitX\\version.txt"
    UNITX_DATA_PATH = "C:\\Users\\UnitX\\unitx_data"
else:
    PRODUCTION_SOFTWARE_VERSION_TXT_PATH = "/home/unitx/prod/production_src/version.txt"
    UNITX_DATA_PATH = "/home/unitx/unitx_data"

logger = formatted_logging.FormattedLogging(__name__).getLog()
def create_benchmark_config(base_benchmark_config: dict, camera_resolution: str, model_resolution: str,
                            seq_interval_ms: int, share_memory_interval_time: int = 40,
                            max_benchmark_time: int = 7200) -> dict:
    camera_resolution_map = {
        '5mp': (2448, 2048),
        '12mp': (4096, 3008),
        '24mp': (5328, 4608),
    }
    camera_width, camera_height = camera_resolution_map[camera_resolution]

    model_resolution_map = {
        '1mp': (960, 803),
        '3mp': (1900, 1589),
        '5mp': (2448, 2048),
        '7mp': (3100, 2276),
        '9mp': (3500, 2570),
        '12mp': (4096, 3008),
        '16mp': (4300, 3718),
        '20mp': (4800, 4151),
        '24mp': (5328, 4608),
    }
    model_width, model_height = model_resolution_map[model_resolution]
    enable_sdk = os.getenv('enabled_sdk', False)
    result_benchmark_config = {
        'seq_interval_ms': seq_interval_ms,
        'is_save_image': True,
        'enable_sdk': enable_sdk,
        'camera_width': camera_width,
        'camera_height': camera_height,
        'camera_resolution': camera_resolution,
        'model_name': f'network_{camera_resolution}_{model_resolution}',
        'model_width': model_width,
        'model_height': model_height,
        'model_resolution': model_resolution,
        'network_architecture': 'v4',
        'ng_type_number': 10,
        'each_ng_type_defect_number': 5,
        'edge_name':'Test'
    }
    # result_benchmark_config.update(base_benchmark_config)
    # result_benchmark_config.update(register_map_func(base_benchmark_config['register_first_twice']))
    result_benchmark_config['max_benchmark_time'] = max_benchmark_time
    result_benchmark_config['share_memory_interval_time'] = share_memory_interval_time

    return result_benchmark_config


def register_map_func(register_first_twice) -> dict:
    return {
        'benchmark_mode': int(f'{register_first_twice}00'),
        'benchmark_trigger_interval_ms': int(f'{register_first_twice}02'),
        'benchmark_reset_interval_ms': int(f'{register_first_twice}88'),
        'benchmark_counter': int(f'{register_first_twice}26'),
        'benchmark_temp_counter': int(f'{register_first_twice}86'),
        'benchmark_mismatch_images_flag': int(f'{register_first_twice}73'),
    }

class DataMonitor(object):
    REPORT_FILE_NAME = "simulation_results.xlsx"
    GENERATE_25D_PATTERN = r"generate_overlaid_image_25d took ([\d.]+) seconds"

    def __init__(self, data_monitor_config):
        self.data_monitor_config = data_monitor_config

        self.wb = None
        self.thread = None
        self.log_start_time = None
        self.stop_system_data_flag = False
        self.cpu_usage = []
        self.gpu_usage = []
        self.gpu_mem_usage = []
        self.memory_usage = []
        self.disk_usage = []
        self.disk_read_speed = []
        self.disk_write_speed = []
        self.all_params = []

        nvmlInit()
        self.handle = nvmlDeviceGetHandleByIndex(0)

        pass

    def __del__(self):
        nvmlShutdown()

    @staticmethod
    def grep_logs_between_start_end(log_file: str, start_pattern: str, end_pattern: str) -> list:
        grep_times = 3
        try:
            for _ in range(grep_times):
                cmd = f"sed -n '/{start_pattern}/,/{end_pattern}/p' {log_file}"
                result = subprocess.run(
                    cmd,
                    shell=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    check=True
                )
                logs_between_start_end = result.stdout
                if logs_between_start_end:
                    return logs_between_start_end.split('\n')
                else:
                    start_pattern = datetime.strptime(start_pattern, '%Y-%m-%d %H:%M:%S').timestamp() + 1
                    start_pattern = datetime.fromtimestamp(start_pattern).strftime('%Y-%m-%d %H:%M:%S')
                    continue
            return []
        except Exception as e:
            logger.error(f"{inspect.currentframe().f_code.co_name} failed {e}")
            return []

    @staticmethod
    def calculate_cortex_infer_time(result_log) -> list:
        """
        获取Cortex推理每张图片的耗时，然后统计平均耗时，并获取最大耗时
        :param result_log:
        :return:
        """
        all_infer_time = []
        for line in result_log:
            if 'cortex batch inference on' in line:
                all_infer_time.append(int(line.split(' ')[-1]))

        if not all_infer_time:
            return [0, 0, 0]
        return [max(all_infer_time), min(all_infer_time), sum(all_infer_time) / len(all_infer_time)]

    @staticmethod
    def get_generate_25d_time_cost(result_log):
        """
        This function is used to get the generate_overlaid_image_25d time cost from log file
        :param result_log:
        :return: a dict {"normal": [max_normal_time, min_normal_time, average_normal_time],
                        "mean": [max_mean_time, min_mean_time, average_mean_time],
                        "height": [max_height_time, min_height_time, average_height_time]}
        """
        all_normal_time = []
        all_mean_time = []
        all_height_time = []

        for line in result_log:
            if 'generate_overlaid_image_25d took' in line:
                matches = re.findall(DataMonitor.GENERATE_25D_PATTERN, line)
                if not matches:
                    logger.error(f"{DataMonitor.GENERATE_25D_PATTERN} not in line {line}")
                    continue

                if "capture_config: NORMAL" in line:
                    all_normal_time.append(matches[0])
                elif "capture_config: MEAN" in line:
                    all_mean_time.append(matches[0])
                elif "capture_config: HEIGHT" in line:
                    all_height_time.append(matches[0])

        if not all_normal_time:
            return_normal_time = [0, 0, 0]
        else:
            return_normal_time = [max(all_normal_time), min(all_normal_time), sum(all_normal_time) / len(all_normal_time)]
        if not all_mean_time:
            return_mean_time = [0, 0, 0]
        else:
            return_mean_time = [max(all_mean_time), min(all_mean_time), sum(all_mean_time) / len(all_mean_time)]
        if not all_height_time:
            return_height_time = [0, 0, 0]
        else:
            return_height_time = [max(all_height_time), min(all_height_time), sum(all_height_time) / len(all_height_time)]

        return {"normal": return_normal_time, "mean": return_mean_time, "height": return_height_time}

    @staticmethod
    def get_int_software_version() -> int:
        with open(PRODUCTION_SOFTWARE_VERSION_TXT_PATH, 'r') as f:
            software_version = f.readlines()[0].strip()
            software_version = software_version.split('-')[0].replace('v', '')
            int_version = int(software_version.split('.')[0]) * 100 + int(software_version.split('.')[1])
        return int_version

    @staticmethod
    def get_current_time():
        return time.strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def get_timestamp_from_log(line: str):
        datetime_str = line.split(' | INFO | ')[0]
        if not datetime_str.startswith('20'):
            datetime_str = datetime_str.split(']')[1]
        try:
            return datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S').timestamp()
        except ValueError:
            return datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S.%f').timestamp()

    @staticmethod
    def get_cpu_info() -> str:
        cpu_info = cpuinfo.get_cpu_info()
        return f"{cpu_info['brand_raw']}"

    @staticmethod
    def get_gpu_info() -> str:
        gpus = GPUtil.getGPUs()
        return str([f"GPU: {gpu.name} {gpu.memoryTotal}MB" for gpu in gpus])

    @staticmethod
    def get_memory_info() -> str:
        """
        If the platform is Windows, it uses the WMI library to get the memory information.
        If the platform is Linux, it uses the subprocess library to run the 'dmidecode --type memory' command.
            You should use 'su admin' and input the password 'test123'.
            Then 'sudo visudo' and add '%unitx ALL=(ALL) NOPASSWD: /usr/sbin/dmidecode' to the end of the file.
            Then save and exit.
        """
        if platform.system() == 'Windows':
            w = wmi.WMI()
            total_memory = 0
            memory_type = None
            memory_speed = 0
            for memory in w.Win32_PhysicalMemory():
                total_memory += int(memory.Capacity)
                memory_speed = memory.Speed
                memory_type = memory.MemoryType
            total_memory_gb = total_memory / (1024 ** 3)
            memory_type_str = {
                20: "DDR",
                21: "DDR2",
                24: "DDR3",
                26: "DDR4",
                25: "DDR5"
            }.get(memory_type, "Unknown")
            return f"{memory_type_str} {memory_speed}MHz {total_memory_gb:.2f}GB"
        else:
            command = "docker exec -u root ${DOCKER_CONTAINER_ID} dmidecode --type memory"
            try:

                process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                                           text=True)

                output, error = process.communicate()
                every_memory = []
                memory_speed = []
                memory_type = []
                memory_info = []
                for line in output.split('\n'):
                    if ('Volatile Size:' not in line and 'Size:' in line) and ('GB' in line or 'MB' in line):
                        size_str = re.search(r'\d+\s*(GB|MB)', line).group()
                        every_memory.append(size_str)
                    if 'Configured Memory Speed:' in line and 'Unknown' not in line:
                        memory_speed.append(re.search(r'\d+', line).group())
                    if 'Type:' in line and 'Unknown' not in line and 'Error' not in line:
                        memory_type.append(line.split(': ')[1].strip())

                for i in range(len(memory_type)):
                    memory_info.append(f"{memory_type[i]} {memory_speed[i]}MT/s {every_memory[i]}")

                return str(memory_info)

            except subprocess.CalledProcessError as e:
                print(f"Error while fetching memory info: {e}")
                return "Information not available"

    @staticmethod
    def get_disk_info() -> str:
        if platform.system() == 'Windows':
            c = wmi.WMI()
            disks = []
            for disk in c.Win32_DiskDrive():
                disks.append(f"{disk.Model} ({int(disk.Size) / (1024 ** 3):.2f}GB)")
            return str(disks)
        else:
            try:
                result = subprocess.run(['lsblk', '-o', 'NAME,MODEL,SIZE'], capture_output=True, text=True, check=True)
                output = result.stdout
                disks = output.split('\n')
                disk_info = []
                for disk in disks:
                    if 'sd' in disk or 'nvme' in disk:
                        if '─' not in disk:
                            disk_info.append(f"{disk.split('    ')[1]}")
                return str(disk_info)
            except FileNotFoundError:
                return "Disk info requires 'lsblk'"

    @staticmethod
    def get_system_info() -> list:
        info = [
            DataMonitor.get_cpu_info(),
            DataMonitor.get_gpu_info(),
            DataMonitor.get_memory_info(),
            DataMonitor.get_disk_info()
        ]
        return info

    @staticmethod
    def plot_and_insert(data, title, cell, ws, data_list):
        safe_title = re.sub(r'[\\/*?:"<>|]', "", title)
        plt.figure(figsize=(14.5, 6))
        plt.plot(data)
        plt.title(title)
        plt.xlabel("Time (500ms)")
        plt.ylabel(title)
        plt.grid(True)
        plt.savefig(f"{safe_title}.png", dpi=100, bbox_inches='tight', pad_inches=0.1)
        plt.close()

        img = Image(f"{safe_title}.png")
        ws.add_image(img, cell)
        data_list.append(f"{safe_title}.png")

    @staticmethod
    def get_process_affinity(process_cmdline_part):
        for proc in psutil.process_iter(['pid', 'name', 'cmdline', 'cpu_affinity']):
            try:
                if proc.info['name'] == 'python3' and any(process_cmdline_part in cmd for cmd in proc.info['cmdline']):
                    return proc.info['cpu_affinity']
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass
        return []

    def calculate_part_time(self, result_log) -> list:
        def get_part_group_id():
            _temp_line_0 = line.split("'group_id': '")[-1]
            return _temp_line_0.split("'")[0]

        start_part_time = {}
        end_part_time = {}
        all_part_time = []
        for line in result_log:
            if 'Starting part group with parts' in line:
                start_part_time[get_part_group_id()] = self.get_timestamp_from_log(line)

            if 'Marking part group as done' in line:
                end_part_time[get_part_group_id()] = self.get_timestamp_from_log(line)

        for key in start_part_time:
            try:
                all_part_time.append(end_part_time[key] - start_part_time[key])
            except KeyError:
                pass
        if not all_part_time:
            return [0, 0, 0]
        logger.info(
            f"all_part_num {len(all_part_time)}, result {[max(all_part_time), min(all_part_time), sum(all_part_time) / len(all_part_time)]}"
        )
        return [max(all_part_time), min(all_part_time), sum(all_part_time) / len(all_part_time)]

    def clear_system_data(self):
        self.stop_system_data_flag = False
        self.cpu_usage = []
        self.gpu_usage = []
        self.gpu_mem_usage = []
        self.memory_usage = []
        self.disk_usage = []
        self.disk_read_speed = []
        self.disk_write_speed = []

    def create_workbook(self):
        try:
            self.wb = load_workbook(self.REPORT_FILE_NAME)
        except FileNotFoundError:
            self.wb = Workbook()

    def get_system_data(self, start_time: float):
        self.log_start_time = datetime.fromtimestamp(start_time).strftime('%Y-%m-%d %H:%M:%S')
        self.thread = Thread(target=self.thread_get_system_data, daemon=True)
        self.thread.start()
        return self.thread

    def stop_system_data(self):
        self.stop_system_data_flag = True
        self.thread.join()

    def thread_get_system_data(self, interval: float = 0.5):
        last_disk_io = psutil.disk_io_counters()
        last_time = time.time()
        self.clear_system_data()

        while not self.stop_system_data_flag:
            self.cpu_usage.append(psutil.cpu_percent(interval=0))

            gpu_util = nvmlDeviceGetUtilizationRates(self.handle)
            self.gpu_usage.append(gpu_util.gpu)
            gpu_mem = nvmlDeviceGetMemoryInfo(self.handle)
            self.gpu_mem_usage.append((gpu_mem.used / gpu_mem.total) * 100)

            memory_info = psutil.virtual_memory()
            self.memory_usage.append(memory_info.percent)

            disk_info = psutil.disk_usage('/')
            self.disk_usage.append(disk_info.percent)

            current_disk_io = psutil.disk_io_counters()
            current_time = time.time()
            read_speed = (current_disk_io.read_bytes - last_disk_io.read_bytes) / (current_time - last_time) / (
                    1024 * 1024)
            write_speed = (current_disk_io.write_bytes - last_disk_io.write_bytes) / (current_time - last_time) / (
                    1024 * 1024)
            self.disk_read_speed.append(read_speed)
            self.disk_write_speed.append(write_speed)
            last_disk_io = current_disk_io
            last_time = current_time

            time.sleep(interval)

    def get_core_allocation(self) -> str:
        processes = {
            "prod service": "production_src/server/run_prod.py",
            "prod ui": "production_src/ui/run_prod.py",
            "cortex": "production_src/server/run_prod_cortex.py",
            "optix": "production_src/server/run_prod_optix.py"
        }
        core_allocation = ''
        for name, cmdline_part in processes.items():
            affinity = self.get_process_affinity(cmdline_part)
            if affinity:
                core_allocation += f"{name}: {', '.join(map(str, affinity))}\n"
            else:
                core_allocation += f"{name}: Process not found\n"
        return core_allocation

    def get_prod_info(self) -> (list, str):
        camera_resolution = self.data_monitor_config["camera_resolution"]
        model_resize = self.data_monitor_config["model_resolution"]
        network_architecture = self.data_monitor_config["network_architecture"]
        with open(PRODUCTION_SOFTWARE_VERSION_TXT_PATH, 'r') as f:
            software_version = f.readlines()[0].strip()
        core_allocation = self.get_core_allocation()
        ng_type_number = self.data_monitor_config["ng_type_number"]
        each_ng_type_defect_number = self.data_monitor_config["each_ng_type_defect_number"]
        is_image_saving = str(self.data_monitor_config["is_save_image"])

        return [camera_resolution, model_resize, network_architecture, software_version, ng_type_number,
                each_ng_type_defect_number, is_image_saving], core_allocation

    def create_report(self, benchmark_data: dict):
        end_time = self.get_current_time()
        result_log = self.grep_logs_between_start_end(PROD_LOG_PATH, self.log_start_time, end_time)
        if "data" in self.wb.sheetnames:
            data_ws = self.wb["data"]
        else:
            data_ws = self.wb.create_sheet(title="data")

        if data_ws.max_row == 1:
            data_ws.append([
                # System Info
                "Edge Name", "Camera Resolution", "Model Resize", "Network Architecture", "Software Version",
                "NG Type Number", 'Each NG Type Defect Number', "Is Image Saving",

                # Benchmark Data
                "Part Count", "Total Use Time(s)", "FPS", "MP/s", "Max Part Use Time (s)", "Min Part Use Time (s)",
                "Avg Part Use Time (s)", "Max Cortex Infer Time (ms)", "Min Cortex Infer Time (ms)",
                "Avg Cortex Infer Time (ms)",

                # System Data
                "CPU", "GPU", "RAM", "SSD", "CPU Usage AVG (%)", "GPU Usage AVG (%)", "GPU Memory Usage AVG (%)",
                "Memory Usage AVG (%)", "Disk Usage AVG (%)", "Disk Read Speed AVG (MB/s)",
                "Disk Write Speed AVG (MB/s)",

                # Core Allocation
                "Core Allocation", "Created At"
            ])
        prod_info, core_allocation = self.get_prod_info()
        mps = int(self.data_monitor_config['model_resolution'].split('mp')[0]) * benchmark_data['fps']
        edge_name = self.data_monitor_config['edge_name']
        data_ws.append([edge_name] + prod_info + list(benchmark_data.values()) + [mps] + self.calculate_part_time(
            result_log) + self.calculate_cortex_infer_time(result_log) + self.get_system_info() +
                       [sum(self.cpu_usage) / len(self.cpu_usage),
                        sum(self.gpu_usage) / len(self.gpu_usage),
                        sum(self.gpu_mem_usage) / len(
                            self.gpu_mem_usage),
                        sum(self.memory_usage) / len(self.memory_usage),
                        sum(self.disk_usage) / len(self.disk_usage),
                        sum(self.disk_read_speed) / len(
                            self.disk_read_speed),
                        sum(self.disk_write_speed) / len(
                            self.disk_write_speed)] + [core_allocation, end_time])
        for column in data_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    logger.error(f"{inspect.currentframe().f_code.co_name} raise error: {e}")
            adjusted_width = (max_length + 2)
            data_ws.column_dimensions[column_letter].width = adjusted_width

        plt_worksheet_name = (f'{edge_name}_{self.data_monitor_config["camera_resolution"]}'
                              f'_{self.data_monitor_config["model_resolution"]}')[:31]
        if plt_worksheet_name in self.wb.sheetnames:
            del self.wb[plt_worksheet_name]
        chart_ws = self.wb.create_sheet(title=plt_worksheet_name)

        self.plot_and_insert(self.cpu_usage, "CPU Usage (%)", "A1", chart_ws, self.all_params)
        self.plot_and_insert(self.gpu_usage, "GPU Usage (%)", "R1", chart_ws, self.all_params)
        self.plot_and_insert(self.gpu_mem_usage, "GPU Memory Usage (%)", "AI1", chart_ws, self.all_params)
        self.plot_and_insert(self.memory_usage, "Memory Usage (%)", "A36", chart_ws, self.all_params)
        self.plot_and_insert(self.disk_usage, "Disk Usage (%)", "R36", chart_ws, self.all_params)
        self.plot_and_insert(self.disk_read_speed, "Disk Read Speed (MB/s)", "AI36", chart_ws, self.all_params)
        self.plot_and_insert(self.disk_write_speed, "Disk Write Speed (MB/s)", "AZ36", chart_ws, self.all_params)

        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        self.wb.save(self.REPORT_FILE_NAME)
        self.all_params.clear()

    def get_image_capture_time_cost(self):
        pass
        # !todo 获取Optix获取每张图片的耗时，然后统计平均耗时，并获取最大耗时










